import pandas as pd
import numpy as np
import sys
import time
import os
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm


# ------------------------------------------------------------
# FLOOR ORDER LOGIC
# ------------------------------------------------------------
def logical_floor_order(floor):
    if pd.isna(floor):
        return np.inf
    floor = str(floor).strip().upper()
    if floor == "G":
        return 0
    try:
        return int(floor)
    except ValueError:
        if "BASE" in floor:
            return -1
        elif "TERRACE" in floor or floor == "T":
            return 100
        else:
            return np.inf


# ------------------------------------------------------------
# COLUMN DETECTION LOGIC
# ------------------------------------------------------------
def detect_column(df, possible_names):
    norm = {c.strip().lower().replace(" ", ""): c for c in df.columns}
    for name in possible_names:
        key = name.lower().replace(" ", "")
        if key in norm:
            return norm[key]
    raise KeyError(f"❌ Cannot find any of: {possible_names}\nColumns available: {list(df.columns)}")


# ------------------------------------------------------------
# SPLITTING LOGIC WITH OPTION B (PROPORTIONAL CARPET SPLIT)
# ------------------------------------------------------------
def process_property(prop_code, area_r, df_prop):

    cumulative = 0
    result_rows = []

    valid_types = ["R", "WR", "SR", "PG", "HO", "ICR"]
    df_valid = df_prop[df_prop["TypeOFUse"].isin(valid_types)].copy()
    df_other = df_prop[~df_prop["TypeOFUse"].isin(valid_types)].copy()

    df_valid = df_valid.sort_values(by="FloorOrder").reset_index(drop=True)

    for idx, row in df_valid.iterrows():

        original_built = float(row["BuiltupAreaSqFeet"])
        original_carpet = float(row["CarpetAreaSqFeet"])

        # Case 1: No split needed for this row
        if cumulative + original_built <= area_r:

            r = row.to_dict()
            r["SplitRow"] = "Balanced Part"
            r["Status"] = "Balanced"

            ratio = r["BuiltupAreaSqFeet"] / original_built
            r["CarpetAreaSqFeet"] = original_carpet * ratio

            result_rows.append(r)
            cumulative += original_built

        else:
            # Split required
            remaining = area_r - cumulative
            overflow = original_built - remaining

            # PART 1 - BALANCED SPLIT
            if remaining > 0:
                part1 = row.copy()
                part1["BuiltupAreaSqFeet"] = remaining
                part1["SplitRow"] = "Balanced Part"
                part1["Status"] = "Balanced"

                ratio = remaining / original_built
                part1["CarpetAreaSqFeet"] = original_carpet * ratio

                result_rows.append(part1.to_dict())

            # PART 2 - EXCESS SPLIT
            part2 = row.copy()
            part2["BuiltupAreaSqFeet"] = overflow
            part2["ConstructionYear"] = 2025
            part2["SplitRow"] = "Overflow Split"
            part2["Status"] = "Excess"

            ratio = overflow / original_built
            part2["CarpetAreaSqFeet"] = original_carpet * ratio

            result_rows.append(part2.to_dict())

            # Remaining floors after this split
            for _, nxt in df_valid.loc[idx + 1:].iterrows():

                original_b = float(nxt["BuiltupAreaSqFeet"])
                original_c = float(nxt["CarpetAreaSqFeet"])

                nxt = nxt.copy()
                nxt["ConstructionYear"] = 2025
                nxt["SplitRow"] = "After Overflow"
                nxt["Status"] = "Excess"

                ratio = 1  # no split for these floors
                nxt["CarpetAreaSqFeet"] = original_c * ratio

                result_rows.append(nxt.to_dict())

            break  # Important: stop after overflow split

    # NON-RESIDENTIAL FLOORS (add as Excess)
    for _, other in df_other.iterrows():
        orig_b = float(other["BuiltupAreaSqFeet"])
        orig_c = float(other["CarpetAreaSqFeet"])

        other = other.copy()
        other["ConstructionYear"] = 2025
        other["SplitRow"] = "Non-Residential"
        other["Status"] = "Excess"

        ratio = 1
        other["CarpetAreaSqFeet"] = orig_c * ratio

        result_rows.append(other.to_dict())

    df_out = pd.DataFrame(result_rows)
    if not df_out.empty:
        df_out["PropertyCode"] = prop_code

    return df_out


# ------------------------------------------------------------
# MAIN SCRIPT
# ------------------------------------------------------------
def main(area_file, floor_file, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    log("\n🏗️ Starting Property Area Split & Proportional Carpet Calculation...\n")
    start = time.time()

    try:
        df_area = pd.read_excel(area_file, engine="openpyxl")
        df_floor = pd.read_excel(floor_file, engine="openpyxl")
    except Exception as e:
        log(f"❌ Error reading files: {e}")
        return

    # Strip spaces
    df_area.columns = df_area.columns.str.strip()
    df_floor.columns = df_floor.columns.str.strip()

    log(f"📘 Area file loaded: {len(df_area)} rows")
    log(f"📗 Floor file loaded: {len(df_floor)} rows\n")

    # Detect columns
    try:
        prop_area = detect_column(df_area, ["PropertyCode"])
        area_col = detect_column(df_area, ["Area_R", "AreaR", "TotalArea"])

        prop_floor = detect_column(df_floor, ["PropertyCode", "propertycode"])
        floor_col = detect_column(df_floor, ["FloorID", "Floor", "Floor Id"])
        builtup_col = detect_column(df_floor, ["BuiltupAreaSqFeet", "BuiltUpArea", "BuiltupAreaSqft"])
        type_col = detect_column(df_floor, ["TypeOFUse", "TypeOfUse"])
        year_col = detect_column(df_floor, ["ConstructionYear", "Year"])
        carpet_col = detect_column(df_floor, ["CarpetAreaSqFeet", "CarpetArea"])
    except KeyError as e:
        log(str(e))
        return

    # Normalize names
    df_area.rename(columns={prop_area: "PropertyCode", area_col: "Area_R"}, inplace=True)
    df_floor.rename(columns={
        prop_floor: "PropertyCode",
        floor_col: "FloorID",
        builtup_col: "BuiltupAreaSqFeet",
        type_col: "TypeOFUse",
        year_col: "ConstructionYear",
        carpet_col: "CarpetAreaSqFeet"
    }, inplace=True)

    # Add sorted floor order
    df_floor["FloorOrder"] = df_floor["FloorID"].apply(logical_floor_order)

    all_results = []

    log(f"🏠 Processing properties...\n")

    # Use tqdm only if no callback, or just log progress periodically
    iterator = df_area.iterrows()
    if not log_callback:
        iterator = tqdm(df_area.iterrows(), total=len(df_area), ncols=90, desc="Processing")
    
    total_props = len(df_area)
    for idx, (index, row) in enumerate(iterator):
        prop = row["PropertyCode"]
        area_r = float(row["Area_R"]) if not pd.isna(row["Area_R"]) else 0
        df_prop = df_floor[df_floor["PropertyCode"] == prop].copy()

        if df_prop.empty or area_r <= 0:
            continue

        total_built = df_prop["BuiltupAreaSqFeet"].sum()
        valid_types = ["R", "WR", "SR", "PG", "HO", "ICR"]

        # Case: No split needed
        if total_built <= area_r:

            df_prop["SplitRow"] = np.where(df_prop["TypeOFUse"].isin(valid_types), "Balanced Part", "Non-Residential")
            df_prop["Status"] = np.where(df_prop["TypeOFUse"].isin(valid_types), "Balanced", "Excess")

            df_prop.loc[~df_prop["TypeOFUse"].isin(valid_types), "ConstructionYear"] = 2025
            df_prop["PropertyCode"] = prop

            all_results.append(df_prop)

        else:
            out = process_property(prop, area_r, df_prop)
            if not out.empty:
                all_results.append(out)
        
        if log_callback and (idx + 1) % 100 == 0:
             log(f"Processed {idx + 1}/{total_props} properties...")

    # Final combined result
    combined = pd.concat(all_results, ignore_index=True)

    # Timestamped output file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.dirname(area_file)
    output_path = os.path.join(output_dir, f"Rvadiv_{timestamp}.xlsx")

    log(f"\n💾 Saving output: {output_path}")

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            combined.to_excel(writer, sheet_name="Combined", index=False)

        # Apply colors
        wb = load_workbook(output_path)
        ws = wb["Combined"]

        yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        headers = [c.value for c in ws[1]]
        split_i = headers.index("SplitRow")
        status_i = headers.index("Status")

        for row in ws.iter_rows(min_row=2):
            split_val = row[split_i].value
            status_val = row[status_i].value

            if split_val in ["Balanced Part", "Overflow Split"]:
                for c in row:
                    c.fill = yellow
            elif status_val == "Excess":
                for c in row:
                    c.fill = red

        wb.save(output_path)
    except Exception as e:
        log(f"❌ Error saving file: {e}")
        return

    log("\n✅ Process Completed Successfully!")
    log(f"Output File: {output_path}")
    log(f"Time Taken: {round(time.time() - start, 2)} seconds\n")
    
    return output_path


# ------------------------------------------------------------
# RUN MAIN
# ------------------------------------------------------------
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: py manage.py <area_file.xlsx> <floor_file.xlsx>")
    else:
        main(sys.argv[1], sys.argv[2])

